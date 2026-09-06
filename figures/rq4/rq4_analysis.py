"""Local, reproducible data preparation and plots for the RQ4 notebook.

Source reports may be CSV sidecars or Excel-only workbooks. They are always read
in place; derived caches and figures are written only to caller-provided paths.
"""

from pathlib import Path
import csv
import hashlib
import json
import os
import re
import time
import warnings
from zipfile import BadZipFile, ZipFile, is_zipfile
import xml.etree.ElementTree as ET

import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import numpy as np
from openpyxl import load_workbook
import pandas as pd

SUFFIXES = {
    "summary": "R_sum__Projet.csv", "files": "Fichiers_-_M_triques.csv",
    "functions": "Function_Size.csv", "issues": "Issues.csv",
    "scattering": "Git_Scattering_-_R_sum_.csv",
    "pairs": "Git_Scattering_-_Paires.csv",
}
SHEETS = {
    "summary": "Résumé Projet",
    "issues": "Issues",
    "files": "Fichiers - Métriques",
    "scattering": "Git Scattering - Résumé",
    "pairs": "Git Scattering - Paires",
    "functions": "Function Size",
}
HEADERS = {
    "summary": ("metric", "value"),
    "issues": ("issue_key", "file", "rule", "rule_source", "severity", "status",
               "line", "message", "assignee", "created_at", "updated_at"),
    "files": ("file", "ncloc", "complexity", "cognitive_complexity", "code_smells",
              "bugs", "vulnerabilities", "duplicated_lines_density",
              "comment_lines_density", "sqale_index", "git_scattering",
              "git_scattering_commits_analyzed", "git_churn_commits",
              "coupling_efferent", "coupling_imports", "ca_afferent_coupling",
              "ca_internal_imports"),
    "scattering": ("type", "path", "git_scattering", "commits_analyzed",
                   "top_coupled_files"),
    "pairs": ("type", "path", "coupled_file", "shared_commits"),
    "functions": ("file", "function", "start_line", "end_line", "size_lines"),
}
CACHE_SCHEMA_VERSION = "xlsx-direct-v1"
ARTIFACT = re.compile(
    r"(?:^|/)(?:vendor|node_modules|\.build|target|dist|build|coverage|third_party|"
    r"legacy_numpy_code|artifacts|generated|cncf-generated)(?:/|$)", re.I
)
TEST_PATH = re.compile(
    r"(?:^|/)(?:tests?|specs?|__tests__|testing|testdata|e2e|conformance)(?:/|$)|"
    r"(?:^|/)(?:test_[^/]+)|(?:_tests?|\.test|\.spec|Tests)\.[^/]+$", re.I
)
METRICS = {
    "ncloc": "Analyzed NCLOC",
    "complexity_per_kloc": "Cyclomatic complexity / KLOC",
    "cognitive_per_kloc": "Cognitive complexity / KLOC",
    "violations_per_kloc": "Violations / KLOC",
    "severe_violations_per_kloc": "Blocker + critical violations / KLOC",
    "debt_per_kloc": "Estimated remediation minutes / KLOC",
    "function_p95": "Function size p95 (lines)",
    "comment_lines_density": "Comment density (%)",
    "duplicated_lines_density": "Duplication density (%)",
}


def _workbook_sheet_names(path):
    """Read only the workbook manifest; do not parse or mutate worksheet data."""
    if not is_zipfile(path):
        return ()
    try:
        with ZipFile(path) as archive:
            root = ET.fromstring(archive.read("xl/workbook.xml"))
        return tuple(node.attrib["name"] for node in root.iter() if node.tag.endswith("sheet"))
    except (BadZipFile, KeyError, ET.ParseError, OSError):
        return ()


def discover(data_root, settle_seconds=30):
    """Inventory CSV or XLSX reports without writing to the extraction directory.

    A report created less than ``settle_seconds`` ago is retained in the manifest but
    is not analyzed on that run. This prevents a concurrently written workbook from
    being mistaken for a corrupt experiment result.
    """
    data_root = Path(data_root)
    records = []
    schemas = []
    now = time.time()
    for directory in sorted(data_root.iterdir()):
        if not directory.is_dir():
            continue
        paths = list(directory.glob("*.xlsx")) + list((directory / "csv").glob("*.csv"))
        by_stem = {}
        pattern = re.compile(rf"^{re.escape(directory.name)}_(\d{{4}}-\d{{2}}-\d{{2}})_([0-9a-f]+)_(.+)$")
        for path in sorted(paths):
            match = pattern.match(path.name)
            if not match:
                raise ValueError(f"Unexpected artifact name: {path}")
            date, sha, suffix = match.groups()
            stem = path.name[:-(len(suffix) + 1)]
            row = by_stem.setdefault(stem, {
                "project": directory.name, "stem": stem,
                "date": date, "sha": sha,
            })
            kind = "workbook" if suffix == "analyse.xlsx" else next(
                (k for k, s in SUFFIXES.items() if s == suffix), None
            )
            if kind is None:
                raise ValueError(f"Unexpected artifact type: {path}")
            row[kind + "_path"] = str(path)
            row[kind + "_bytes"] = path.stat().st_size
            row[kind + "_mtime_ns"] = path.stat().st_mtime_ns
            if kind != "workbook":
                with path.open(encoding="utf-8-sig", newline="") as handle:
                    header = next(csv.reader(handle), [])
                schemas.append({"source": "csv", "kind": kind,
                                "header": ", ".join(header)})
        for row in by_stem.values():
            workbook = row.get("workbook_path")
            if workbook:
                sheet_names = _workbook_sheet_names(Path(workbook))
                row["workbook_valid"] = bool(sheet_names)
                row["workbook_complete"] = set(SHEETS.values()).issubset(sheet_names)
                row["workbook_settled"] = (
                    now - Path(workbook).stat().st_mtime >= settle_seconds
                )
        records.extend(by_stem.values())
    if not records:
        raise ValueError(f"No RQ4 extraction artifacts found in {data_root}")
    manifest = pd.DataFrame(records)
    for kind in [*SUFFIXES, "workbook"]:
        for field in ("path", "bytes", "mtime_ns"):
            col = kind + "_" + field
            if col not in manifest:
                manifest[col] = np.nan
    for col in ("workbook_valid", "workbook_complete", "workbook_settled"):
        if col not in manifest:
            manifest[col] = False
        manifest[col] = manifest[col].fillna(False).astype(bool)
    manifest["date"] = pd.to_datetime(manifest["date"])
    csv_summary = manifest["summary_bytes"].fillna(0).gt(0)
    manifest["complete_csv"] = manifest[[k + "_bytes" for k in SUFFIXES]].fillna(0).gt(0).all(axis=1)
    manifest["zero_byte_workbook"] = manifest["workbook_bytes"].eq(0)
    manifest["unsettled_workbook"] = manifest["workbook_bytes"].fillna(0).gt(0) & ~manifest["workbook_settled"]
    manifest["workbook_ready"] = manifest["workbook_complete"] & manifest["workbook_settled"]
    manifest["has_summary"] = csv_summary | manifest["workbook_ready"]
    manifest["complete_tables"] = manifest["complete_csv"] | manifest["workbook_ready"]
    manifest["source_format"] = np.select(
        [manifest["complete_csv"], manifest["workbook_ready"]],
        ["csv", "xlsx"], default="unavailable",
    )
    manifest["snapshot_id"] = manifest["project"] + ":" + manifest["stem"]
    schema_rows = []
    if schemas:
        schema_rows.extend(pd.DataFrame(schemas).value_counts().rename("files").reset_index().to_dict("records"))
    workbook_count = int(manifest["workbook_complete"].sum())
    if workbook_count:
        schema_rows.extend({
            "source": "xlsx (expected; validated on load)", "kind": kind,
            "header": ", ".join(HEADERS[kind]),
            "files": workbook_count,
        } for kind in SHEETS)
    schema = pd.DataFrame(schema_rows, columns=["source", "kind", "header", "files"])
    return manifest.sort_values(["project", "date", "sha"]).reset_index(drop=True), schema


def read_csv(path):
    if pd.isna(path) or Path(path).stat().st_size == 0:
        return None
    return pd.read_csv(path)


def _record_value(record, field):
    if isinstance(record, pd.Series) or isinstance(record, dict):
        return record.get(field, np.nan)
    return getattr(record, field, np.nan)


def _worksheet_frame(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    header = tuple(next(rows, ()))
    return pd.DataFrame(rows, columns=header)


def _validate_table(table, kind, source):
    if table is None:
        return None
    actual = tuple(table.columns)
    if actual != HEADERS[kind]:
        raise ValueError(
            f"Unexpected {kind} schema in {source}: expected {HEADERS[kind]}, got {actual}"
        )
    return table


def read_artifact(record, kind, workbook=None):
    """Read an artifact from CSV when available, otherwise from its XLSX sheet."""
    csv_path = _record_value(record, kind + "_path")
    if pd.notna(csv_path) and Path(csv_path).is_file() and Path(csv_path).stat().st_size:
        return _validate_table(read_csv(csv_path), kind, csv_path)
    workbook_path = _record_value(record, "workbook_path")
    if pd.isna(workbook_path) or not Path(workbook_path).is_file():
        return None
    if workbook is not None:
        return _validate_table(_worksheet_frame(workbook[SHEETS[kind]]), kind, workbook_path)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        opened = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return _validate_table(_worksheet_frame(opened[SHEETS[kind]]), kind, workbook_path)
    finally:
        opened.close()


def ratios(frame):
    def column(name, default=np.nan):
        return frame[name] if name in frame else pd.Series(default, index=frame.index, dtype=float)

    for numerator, name in [
        ("complexity", "complexity"), ("cognitive_complexity", "cognitive"),
        ("code_smells", "smells"), ("violations", "violations"),
        ("sqale_index", "debt"), ("bugs", "bugs"), ("vulnerabilities", "vulnerabilities"),
    ]:
        frame[name + "_per_kloc"] = frame[numerator] / frame["ncloc"].where(frame["ncloc"] > 0) * 1000
    frame["source_complexity_per_kloc"] = frame["source_complexity"] / frame["source_ncloc"].where(frame["source_ncloc"] > 0) * 1000
    frame["test_path_ncloc_share"] = frame["test_path_ncloc"] / frame["file_ncloc"].where(frame["file_ncloc"] > 0) * 100
    frame["artifact_ncloc_share"] = (frame["file_ncloc"] - frame["source_ncloc"]) / frame["file_ncloc"].where(frame["file_ncloc"] > 0) * 100
    frame["churn_lines_per_history_commit"] = frame["git_churn_lines"] / frame["git_churn_commits"].where(frame["git_churn_commits"] > 0)
    blocker = pd.to_numeric(
        frame["blocker_violations"] if "blocker_violations" in frame else pd.Series(0, index=frame.index),
        errors="coerce",
    ).fillna(0)
    critical = pd.to_numeric(
        frame["critical_violations"] if "critical_violations" in frame else pd.Series(0, index=frame.index),
        errors="coerce",
    ).fillna(0)
    frame["severe_violations"] = blocker + critical
    frame["severe_violations_per_kloc"] = frame["severe_violations"] / frame["ncloc"].where(frame["ncloc"] > 0) * 1000
    frame["security_hotspots_per_kloc"] = column("security_hotspots") / frame["ncloc"].where(frame["ncloc"] > 0) * 1000
    frame["file_summary_ncloc_ratio"] = frame["file_ncloc"] / frame["ncloc"].where(frame["ncloc"] > 0)
    functions = column("functions")
    detailed = column("detailed_issue_rows")
    frame["function_summary_ratio"] = column("function_rows") / functions.where(functions > 0)
    frame["issue_summary_ratio"] = detailed / frame["violations"].where(frame["violations"] > 0)
    comparable = detailed.notna() & frame["violations"].notna()
    frame["issue_count_matches_summary"] = np.where(
        comparable, detailed.eq(frame["violations"]).astype(float), np.nan
    )
    return frame


def _input_fingerprint(row):
    """Fingerprint the exact manifest state selected for one snapshot."""
    parts = [str(_record_value(row, "source_format"))]
    if _record_value(row, "source_format") == "csv":
        kinds = SUFFIXES
    else:
        kinds = ("workbook",)
    for kind in kinds:
        parts.extend([
            str(_record_value(row, kind + "_path")),
            str(_record_value(row, kind + "_bytes")),
            str(_record_value(row, kind + "_mtime_ns")),
        ])
    return hashlib.sha256("|".join(parts).encode()).hexdigest()[:24]


def load_snapshots(manifest, cache_dir, refresh=False):
    """Aggregate snapshots from CSV or XLSX using an incremental read-only cache.

    The cache lives under ``cache_dir``. New experiment reports only require new rows;
    unchanged reports are reused. A workbook whose size or mtime changes while it is
    being read is skipped and retried on the next run.
    """
    cache_dir = Path(cache_dir).resolve()
    source_parents = []
    for column in ["workbook_path", *[kind + "_path" for kind in SUFFIXES]]:
        if column in manifest:
            source_parents.extend(
                str(Path(value).resolve().parent) for value in manifest[column].dropna()
            )
    if source_parents:
        source_root = Path(os.path.commonpath(source_parents))
        if cache_dir.is_relative_to(source_root):
            raise ValueError(f"Cache directory must be outside read-only report tree: {source_root}")
    analysis_version = hashlib.sha256(
        f"{CACHE_SCHEMA_VERSION}:{pd.__version__}:{np.__version__}".encode()
    ).hexdigest()[:20]
    cache_dir.mkdir(parents=True, exist_ok=True)
    cached = cache_dir / "snapshots-incremental.csv"
    reusable = {}
    if cached.exists() and not refresh:
        try:
            old = pd.read_csv(cached, low_memory=False)
            if {"snapshot_id", "_input_fingerprint", "_analysis_version"}.issubset(old):
                old = old[old["_analysis_version"].eq(analysis_version)]
                reusable = {
                    (row["snapshot_id"], row["_input_fingerprint"]): row
                    for row in old.to_dict("records")
                }
        except (OSError, pd.errors.ParserError) as error:
            warnings.warn(f"Ignoring unreadable snapshot cache {cached}: {error}")
    result = []
    selected = manifest[manifest["has_summary"]]
    reused = processed = 0
    skipped = []
    for i, row in enumerate(selected.itertuples(index=False), 1):
        fingerprint = _input_fingerprint(row)
        key = (row.snapshot_id, fingerprint)
        if key in reusable:
            result.append(reusable[key])
            reused += 1
            continue
        workbook = None
        before = None
        try:
            if row.source_format == "xlsx":
                workbook_path = Path(row.workbook_path)
                before = workbook_path.stat()
                if before.st_size != row.workbook_bytes or before.st_mtime_ns != row.workbook_mtime_ns:
                    raise RuntimeError("workbook changed after discovery")
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", message="Workbook contains no default style")
                    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            summary = read_artifact(row, "summary", workbook=workbook)
            if summary is None or not set(HEADERS["summary"]).issubset(summary.columns):
                raise ValueError("missing or invalid project-summary sheet")
            files = read_artifact(row, "files", workbook=workbook)
            functions = read_artifact(row, "functions", workbook=workbook)
            if row.source_format == "xlsx":
                issue_sheet = workbook[SHEETS["issues"]]
                issue_header = tuple(next(issue_sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()))
                if issue_header != HEADERS["issues"]:
                    raise ValueError(f"unexpected issues schema: {issue_header}")
                detailed_issue_rows = max((issue_sheet.max_row or 1) - 1, 0)
            else:
                issues = read_artifact(row, "issues")
                detailed_issue_rows = len(issues) if issues is not None else np.nan
        except (BadZipFile, KeyError, OSError, RuntimeError, ValueError) as error:
            skipped.append((row.snapshot_id, str(error)))
            if workbook is not None:
                workbook.close()
            continue
        finally:
            if workbook is not None:
                workbook.close()
        if row.source_format == "xlsx":
            after = Path(row.workbook_path).stat()
            if after.st_size != before.st_size or after.st_mtime_ns != before.st_mtime_ns:
                skipped.append((row.snapshot_id, "workbook changed while being read"))
                continue
        if summary["metric"].duplicated().any():
            raise ValueError(f"Duplicate metric in {row.summary_path}")
        record = summary.set_index("metric")["value"].to_dict()
        record.update(
            project=row.project, date=row.date, sha=row.sha, stem=row.stem,
            snapshot_id=row.snapshot_id, input_format=row.source_format,
            _input_fingerprint=fingerprint, _analysis_version=analysis_version,
        )
        record.update(file_rows=np.nan, file_ncloc=np.nan, source_ncloc=np.nan,
                      source_complexity=np.nan, file_size_p95=np.nan, test_path_ncloc=np.nan,
                      top10_complexity_share=np.nan)
        if files is not None:
            for col in files.columns.drop("file"):
                files[col] = pd.to_numeric(files[col], errors="coerce")
            source = files[~files["file"].str.contains(ARTIFACT, na=False)]
            ncloc = files["ncloc"].sum(min_count=1)
            complexity = files["complexity"].sum(min_count=1)
            record.update(
                file_rows=len(files), file_ncloc=ncloc,
                source_ncloc=source["ncloc"].sum(min_count=1) if len(files) else np.nan,
                source_complexity=source["complexity"].sum(min_count=1) if len(files) else np.nan,
                file_size_p95=files.loc[files["ncloc"] > 0, "ncloc"].quantile(.95),
                test_path_ncloc=files.loc[files["file"].str.contains(TEST_PATH, na=False), "ncloc"].sum() if len(files) else np.nan,
                top10_complexity_share=100 * files["complexity"].nlargest(10).sum() / complexity if complexity > 0 else np.nan,
            )
        record.update(function_rows=np.nan, function_median=np.nan, function_p95=np.nan,
                      function_p99=np.nan, function_share_gt100=np.nan, source_function_p95=np.nan)
        if functions is not None:
            sizes = pd.to_numeric(functions["size_lines"], errors="coerce").dropna()
            sizes = sizes[sizes > 0]
            source_sizes = pd.to_numeric(functions.loc[
                ~functions["file"].str.contains(ARTIFACT, na=False), "size_lines"
            ], errors="coerce")
            source_sizes = source_sizes[source_sizes > 0]
            record.update(function_rows=len(functions), function_median=sizes.median(),
                          function_p95=sizes.quantile(.95), function_p99=sizes.quantile(.99),
                          function_share_gt100=100 * (sizes > 100).mean() if len(sizes) else np.nan,
                          source_function_p95=source_sizes.quantile(.95))
        record["detailed_issue_rows"] = detailed_issue_rows
        result.append(record)
        processed += 1
        if processed % 250 == 0:
            print(f"Aggregated {processed:,} new/changed snapshots; reused {reused:,}", flush=True)
    if not result:
        raise ValueError("No stable, usable snapshot reports could be loaded")
    frame = pd.DataFrame(result)
    textual = {"project", "date", "sha", "stem", "snapshot_id", "input_format",
               "_input_fingerprint", "_analysis_version", "alert_status",
               "ncloc_language_distribution", "quality_profiles", "quality_gate_details",
               "maintainability_issues", "reliability_issues", "security_issues",
               "analysis_from_sonarqube_9_4", "git_scattering_top_coupled_files"}
    numeric_columns = frame.columns.difference(list(textual))
    frame = pd.concat([
        frame.drop(columns=numeric_columns),
        frame[numeric_columns].apply(pd.to_numeric, errors="coerce"),
    ], axis=1)
    frame["date"] = pd.to_datetime(frame["date"])
    frame["sonar_date"] = pd.to_datetime(frame["last_commit_date"], unit="ms", errors="coerce", utc=True).dt.tz_localize(None).dt.normalize()
    frame["date_mismatch"] = frame["sonar_date"].ne(frame["date"]) & frame["sonar_date"].notna()
    frame = ratios(frame)
    temporary = cached.with_suffix(".tmp")
    frame.to_csv(temporary, index=False)
    temporary.replace(cached)
    print(f"Loaded {len(frame):,} snapshots ({processed:,} processed, {reused:,} reused, {len(skipped):,} skipped).")
    if skipped:
        warnings.warn(f"Skipped {len(skipped)} unstable/invalid reports; first examples: {skipped[:3]}")
    return frame


def daily_panel(snapshots):
    """One observation per project/calendar day; no guessed SHA ordering."""
    numeric = snapshots.select_dtypes(include="number").columns.tolist()
    panel = snapshots.groupby(["project", "date"])[numeric].median().reset_index()
    counts = snapshots.groupby(["project", "date"]).size().rename("snapshots_in_day")
    return panel.merge(counts, on=["project", "date"], validate="one_to_one")


def project_inventory(manifest, snapshots):
    inventory = manifest.groupby("project").agg(
        artifact_snapshots=("stem", "size"), usable_snapshots=("has_summary", "sum"),
        complete_csv_snapshots=("complete_csv", "sum"), zero_byte_workbooks=("zero_byte_workbook", "sum"),
        complete_table_snapshots=("complete_tables", "sum"),
        valid_workbooks=("workbook_complete", "sum"),
        unsettled_workbooks=("unsettled_workbook", "sum"),
        first_artifact_date=("date", "min"), last_artifact_date=("date", "max"),
    )
    observed = snapshots.groupby("project").agg(
        observed_days=("date", "nunique"), date_mismatches=("date_mismatch", "sum"),
        issue_snapshots=("detailed_issue_rows", lambda s: (s > 0).sum()),
        issue_count_matches=("issue_count_matches_summary", lambda s: s.sum(min_count=1)),
        function_snapshots=("function_rows", lambda s: (s > 0).sum()),
        zero_ncloc_snapshots=("ncloc", lambda s: (s == 0).sum()),
    )
    return inventory.join(observed).sort_index()


def monthly_panels(daily):
    monthly = daily.assign(month=daily["date"].dt.to_period("M").dt.to_timestamp())
    monthly = monthly.groupby(["project", "month"])[list(METRICS)].median().reset_index()
    # Pair the same project in adjacent calendar months, not arbitrary observations.
    previous = monthly.copy()
    previous["month"] = previous["month"] + pd.offsets.MonthBegin(1)
    paired = monthly.merge(previous, on=["project", "month"], suffixes=("", "_previous"))
    for metric in METRICS:
        paired[metric + "_delta"] = paired[metric] - paired[metric + "_previous"]
    return monthly, paired


def endpoint_table(daily):
    rows = []
    for project, group in daily.groupby("project"):
        group = group.sort_values("date")
        for metric in METRICS:
            valid = group[["date", metric]].dropna()
            if valid.empty:
                continue
            first, last = valid.iloc[0], valid.iloc[-1]
            enough = len(valid) >= 2
            rows.append({"project": project, "metric": metric, "days": len(valid),
                         "first_date": first["date"], "last_date": last["date"],
                         "first": first[metric], "last": last[metric],
                         "delta": last[metric] - first[metric] if enough else np.nan,
                         "relative_change_pct": 100 * (last[metric] - first[metric]) / abs(first[metric]) if enough and first[metric] != 0 else np.nan,
                         "time_rank_rho": valid["date"].rank().corr(valid[metric].rank()) if len(valid) >= 3 and valid[metric].nunique() > 1 else np.nan})
    return pd.DataFrame(rows)


def latest_details(manifest, snapshots):
    """Analyze an actual snapshot only when the latest filename day is unique."""
    latest_day = snapshots.groupby("project")["date"].transform("max")
    candidates = snapshots[snapshots["date"].eq(latest_day)]
    ambiguous = candidates.groupby("project").size().loc[lambda s: s > 1]
    selected = candidates[~candidates["project"].isin(ambiguous.index)]
    issues, hotspots, languages, gates = [], [], [], []
    paths = manifest.set_index("snapshot_id")
    for row in selected.itertuples(index=False):
        record = paths.loc[row.snapshot_id]
        detail = read_artifact(record, "issues")
        if detail is not None and not detail.empty:
            detail["project"] = row.project
            issues.append(detail)
        files = read_artifact(record, "files")
        if files is not None and not files.empty:
            files["cognitive_complexity"] = pd.to_numeric(files["cognitive_complexity"], errors="coerce")
            files = files[~files["file"].str.contains(ARTIFACT, na=False)].nlargest(5, "cognitive_complexity")
            files["project"] = row.project
            hotspots.append(files)
        for part in str(row.ncloc_language_distribution).split(";"):
            if "=" in part:
                language, value = part.split("=", 1)
                if language != "<null>" and float(value) > 0:
                    languages.append({"project": row.project, "language": language, "ncloc": float(value)})
        try:
            gate = json.loads(row.quality_gate_details)
        except (TypeError, json.JSONDecodeError):
            gate = {}
        for condition in gate.get("conditions", []):
            if condition.get("level") == "ERROR" or condition.get("status") == "ERROR":
                gates.append({"project": row.project, **condition})
    return {
        "ambiguous_latest": ambiguous, "selected": selected,
        "issues": pd.concat(issues, ignore_index=True) if issues else pd.DataFrame(),
        "hotspots": pd.concat(hotspots, ignore_index=True) if hotspots else pd.DataFrame(),
        "languages": pd.DataFrame(languages), "gate_failures": pd.DataFrame(gates),
    }


def save_figure(fig, output_dir, name):
    fig.savefig(output_dir / f"{name}.pdf", bbox_inches="tight")


def trajectory_pages(daily, output_dir, show=True, per_page=12):
    projects = sorted(daily["project"].unique())
    with PdfPages(output_dir / "all_project_trajectories.pdf") as pdf:
        for offset in range(0, len(projects), per_page):
            subset = projects[offset:offset + per_page]
            fig, axes = plt.subplots(int(np.ceil(len(subset) / 3)), 3, figsize=(15, 3 * int(np.ceil(len(subset) / 3))), squeeze=False)
            for ax, project in zip(axes.flat, subset):
                group = daily[daily["project"] == project].sort_values("date")
                ax.plot(group["date"], group["complexity_per_kloc"], ".-", lw=.8, ms=2, label="all analyzed")
                ax.plot(group["date"], group["source_complexity_per_kloc"], "--", lw=1, label="artifact-filtered")
                ax.set_title(project, fontsize=8, wrap=True)
                ax.set_ylabel("Complexity / KLOC", fontsize=8)
                ax.tick_params(axis="both", labelsize=7)
                ax.tick_params(axis="x", rotation=30)
                if group["complexity_per_kloc"].notna().sum() == 0:
                    ax.text(.5, .5, "No positive NCLOC denominator", ha="center", transform=ax.transAxes, fontsize=8)
            for ax in list(axes.flat)[len(subset):]:
                ax.set_visible(False)
            axes.flat[0].legend(fontsize=7)
            fig.suptitle(f"Every project: daily median complexity density ({offset + 1}–{offset + len(subset)} of {len(projects)})", fontsize=12)
            fig.tight_layout()
            pdf.savefig(fig)
            if show:
                plt.show()
            plt.close(fig)


def project_profile(project, daily, snapshots, details, output_dir):
    group = daily[daily["project"] == project].sort_values("date")
    if group.empty:
        print(f"{project}: no usable summary snapshots.")
        return
    panels = [
        (["ncloc"], "Analyzed NCLOC"),
        (["complexity_per_kloc", "source_complexity_per_kloc"], "Complexity / KLOC"),
        (["function_p95", "source_function_p95", "file_size_p95"], "Size tails (lines)"),
        (["violations_per_kloc", "smells_per_kloc", "severe_violations_per_kloc",
          "security_hotspots_per_kloc"], "Issues / KLOC"),
        (["comment_lines_density", "duplicated_lines_density", "test_path_ncloc_share", "artifact_ncloc_share"], "Percent (distinct denominators)"),
        (["coupling_efferent", "ca_afferent_coupling"], "Exported dependency counts"),
    ]
    fig, axes = plt.subplots(2, 3, figsize=(15, 8))
    for ax, (columns, label) in zip(axes.flat, panels):
        for col in columns:
            ax.plot(group["date"], group[col], ".-", ms=3, label=col)
        ax.set_ylabel(label)
        ax.tick_params(axis="x", rotation=30)
        ax.legend(fontsize=7)
    fig.suptitle(f"{project} — {len(group)} observed calendar days", fontsize=13)
    fig.tight_layout()
    save_figure(fig, output_dir, f"profile-{project}")
    plt.show()
    plt.close(fig)


def history_profile(project, daily, manifest, details, output_dir):
    """History series and bounded drill-down; never pool pair rows across snapshots."""
    group = daily[daily["project"] == project].sort_values("date")
    if group.empty:
        return {}
    fig, axes = plt.subplots(1, 2, figsize=(12, 3.5))
    for ax, metric, title in [
        (axes[0], "git_scattering", "Git scattering (exported history scale)"),
        (axes[1], "churn_lines_per_history_commit", "History churn lines / analyzed history commit"),
    ]:
        ax.plot(group["date"], group[metric], ".-", ms=3)
        ax.set_title(title)
        ax.tick_params(axis="x", rotation=30)
    fig.suptitle(f"{project} — history proxies, not sampled-commit diffs")
    fig.tight_layout()
    save_figure(fig, output_dir, f"history-{project}")
    plt.show()
    plt.close(fig)
    selected = details["selected"].loc[lambda d: d["project"].eq(project)]
    if selected.empty:
        return {}
    record = manifest.set_index("snapshot_id").loc[selected.iloc[0]["snapshot_id"]]
    result = {}
    for kind, metric in [("scattering", "git_scattering"), ("pairs", "shared_commits")]:
        table = read_artifact(record, kind)
        if table is not None and not table.empty:
            table[metric] = pd.to_numeric(table[metric], errors="coerce")
            # Keep only file entities: directory/project rows can overlap them.
            source = table[table["type"].str.lower().eq("file") & ~table["path"].str.contains(ARTIFACT, na=False)]
            if "coupled_file" in source:
                source = source[~source["coupled_file"].str.contains(ARTIFACT, na=False)]
            result[kind] = source.nlargest(10, metric)
    return result
